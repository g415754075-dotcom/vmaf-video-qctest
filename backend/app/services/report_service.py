"""报告生成服务"""
import json
import secrets
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional, Dict, Any
from io import BytesIO

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference, ScatterChart as XLScatterChart
from openpyxl.chart.series import XYSeries
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from app.core.config import settings
from app.models.video import Assessment, Report, TaskStatus, Video
from app.services.assessment_service import assessment_service


def get_quality_rating(vmaf: float) -> Dict[str, str]:
    """根据 VMAF 分数获取质量评级信息"""
    if vmaf > 93:
        return {
            "stars": "★★★★★",
            "level": "优秀",
            "description": "画质非常清晰，几乎无损",
            "recommendation": "强烈推荐",
            "recommendation_icon": "🏆",
            "color": "#22c55e"  # green
        }
    elif vmaf > 85:
        return {
            "stars": "★★★★☆",
            "level": "良好",
            "description": "画质清晰，轻微损失",
            "recommendation": "推荐",
            "recommendation_icon": "✅",
            "color": "#84cc16"  # lime
        }
    elif vmaf > 70:
        return {
            "stars": "★★★☆☆",
            "level": "可接受",
            "description": "画质一般，有明显损失",
            "recommendation": "可用",
            "recommendation_icon": "⚠️",
            "color": "#eab308"  # yellow
        }
    elif vmaf > 50:
        return {
            "stars": "★★☆☆☆",
            "level": "较差",
            "description": "画质模糊，损失较大",
            "recommendation": "不推荐",
            "recommendation_icon": "❌",
            "color": "#f97316"  # orange
        }
    else:
        return {
            "stars": "★☆☆☆☆",
            "level": "很差",
            "description": "画质很差，严重失真",
            "recommendation": "避免使用",
            "recommendation_icon": "🚫",
            "color": "#ef4444"  # red
        }


def calculate_efficiency(vmaf: float, bitrate_mbps: float) -> Dict[str, Any]:
    """计算码率效率"""
    if bitrate_mbps <= 0:
        return {"value": 0, "level": "未知", "description": "无法计算"}

    efficiency = vmaf / bitrate_mbps

    if efficiency > 30:
        return {"value": efficiency, "level": "非常高", "description": "极高性价比"}
    elif efficiency > 20:
        return {"value": efficiency, "level": "高", "description": "高性价比"}
    elif efficiency > 12:
        return {"value": efficiency, "level": "中等", "description": "性价比一般"}
    elif efficiency > 6:
        return {"value": efficiency, "level": "低", "description": "性价比较低"}
    else:
        return {"value": efficiency, "level": "很低", "description": "性价比很低"}


class ReportService:
    """报告生成服务类"""

    def __init__(self):
        self.reports_dir = settings.reports_dir

    async def create_report(
        self,
        session: AsyncSession,
        name: str,
        assessment_ids: List[int],
        include_sections: List[str]
    ) -> Report:
        """创建报告记录"""
        # 验证评估任务存在且已完成
        for aid in assessment_ids:
            assessment = await session.get(Assessment, aid)
            if not assessment:
                raise ValueError(f"评估任务 {aid} 不存在")
            if assessment.status != TaskStatus.COMPLETED:
                raise ValueError(f"评估任务 {aid} 尚未完成")

        # 确定报告类型
        report_type = "single" if len(assessment_ids) == 1 else "comparison"

        # 创建报告记录
        report = Report(
            name=name,
            report_type=report_type,
            assessment_ids={"ids": assessment_ids, "sections": include_sections}
        )

        session.add(report)
        await session.commit()
        await session.refresh(report)

        # 生成报告文件
        await self._generate_report_files(session, report)

        return report

    async def _generate_report_files(
        self,
        session: AsyncSession,
        report: Report
    ) -> None:
        """生成报告文件（PDF、Excel、JSON）"""
        report_dir = self.reports_dir / f"report_{report.id}"
        report_dir.mkdir(parents=True, exist_ok=True)

        assessment_ids = report.assessment_ids.get("ids", [])
        sections = report.assessment_ids.get("sections", [])

        # 获取评估数据
        assessments_data = []
        for aid in assessment_ids:
            # 使用 select 语句预加载关联的视频对象
            query = (
                select(Assessment)
                .where(Assessment.id == aid)
                .options(
                    selectinload(Assessment.reference_video),
                    selectinload(Assessment.distorted_video)
                )
            )
            result = await session.execute(query)
            assessment = result.scalar_one_or_none()
            if assessment:
                frame_data = await assessment_service.get_frame_data(session, aid)
                stats = await assessment_service.get_statistics(session, aid)
                assessments_data.append({
                    "assessment": assessment,
                    "frame_data": frame_data,
                    "statistics": stats
                })

        # 生成 JSON
        json_path = report_dir / "report.json"
        await self._generate_json(json_path, assessments_data)
        report.json_path = str(json_path)

        # 生成 Excel
        excel_path = report_dir / "report.xlsx"
        await self._generate_excel(excel_path, assessments_data, sections)
        report.excel_path = str(excel_path)

        # 生成 PDF
        pdf_path = report_dir / "report.pdf"
        await self._generate_pdf(pdf_path, assessments_data, sections, report.name)
        report.pdf_path = str(pdf_path)

        await session.commit()

    async def _generate_json(
        self,
        output_path: Path,
        assessments_data: List[Dict]
    ) -> None:
        """生成 JSON 报告"""
        report_data = {
            "generated_at": datetime.utcnow().isoformat(),
            "assessments": []
        }

        for data in assessments_data:
            assessment = data["assessment"]
            report_data["assessments"].append({
                "id": assessment.id,
                "reference_video": {
                    "filename": assessment.reference_video.original_filename,
                    "resolution": f"{assessment.reference_video.width}x{assessment.reference_video.height}",
                    "codec": assessment.reference_video.codec
                },
                "distorted_video": {
                    "filename": assessment.distorted_video.original_filename,
                    "resolution": f"{assessment.distorted_video.width}x{assessment.distorted_video.height}",
                    "codec": assessment.distorted_video.codec,
                    "bitrate": assessment.distorted_video.bitrate
                },
                "scores": {
                    "vmaf": assessment.vmaf_score,
                    "vmaf_min": assessment.vmaf_min,
                    "vmaf_max": assessment.vmaf_max,
                    "ssim": assessment.ssim_score,
                    "psnr": assessment.psnr_score
                },
                "statistics": data["statistics"],
                "frame_data": data["frame_data"]
            })

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)

    async def _generate_excel(
        self,
        output_path: Path,
        assessments_data: List[Dict],
        sections: List[str]
    ) -> None:
        """生成 Excel 报告"""
        wb = Workbook()

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # 摘要 Sheet
        ws_summary = wb.active
        ws_summary.title = "摘要"

        summary_headers = ["视频名称", "分辨率", "编码器", "码率(Mbps)", "VMAF", "SSIM", "PSNR", "质量等级"]
        ws_summary.append(summary_headers)

        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for data in assessments_data:
            assessment = data["assessment"]
            dist_video = assessment.distorted_video

            # 质量等级判断
            vmaf = assessment.vmaf_score or 0
            if vmaf > 90:
                quality_level = "优秀"
            elif vmaf > 80:
                quality_level = "良好"
            elif vmaf > 70:
                quality_level = "可接受"
            else:
                quality_level = "差"

            bitrate_mbps = (dist_video.bitrate or 0) / 1_000_000

            row = [
                dist_video.original_filename,
                f"{dist_video.width}x{dist_video.height}",
                dist_video.codec or "N/A",
                f"{bitrate_mbps:.2f}",
                f"{assessment.vmaf_score:.2f}" if assessment.vmaf_score else "N/A",
                f"{assessment.ssim_score:.4f}" if assessment.ssim_score else "N/A",
                f"{assessment.psnr_score:.2f}" if assessment.psnr_score else "N/A",
                quality_level
            ]
            ws_summary.append(row)

        # 调整列宽
        for col in ws_summary.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_summary.column_dimensions[col[0].column_letter].width = max_length + 2

        # 逐帧数据 Sheet
        if "charts" in sections or "statistics" in sections:
            for i, data in enumerate(assessments_data):
                assessment = data["assessment"]
                frame_data = data["frame_data"] or []

                ws_frames = wb.create_sheet(title=f"逐帧数据_{i+1}")

                frame_headers = ["帧号", "VMAF", "SSIM", "PSNR"]
                ws_frames.append(frame_headers)

                for col, header in enumerate(frame_headers, 1):
                    cell = ws_frames.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill

                for frame in frame_data:
                    ws_frames.append([
                        frame.get("frame_num", 0),
                        frame.get("vmaf"),
                        frame.get("ssim"),
                        frame.get("psnr")
                    ])

                # 添加图表
                if len(frame_data) > 0:
                    chart = LineChart()
                    chart.title = "VMAF 质量曲线"
                    chart.x_axis.title = "帧号"
                    chart.y_axis.title = "VMAF"
                    chart.y_axis.scaling.min = 0
                    chart.y_axis.scaling.max = 100

                    data_ref = Reference(ws_frames, min_col=2, min_row=1, max_row=len(frame_data)+1)
                    categories = Reference(ws_frames, min_col=1, min_row=2, max_row=len(frame_data)+1)

                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(categories)
                    chart.width = 20
                    chart.height = 10

                    ws_frames.add_chart(chart, "F2")

        # 统计数据 Sheet
        if "statistics" in sections:
            ws_stats = wb.create_sheet(title="统计分析")

            stats_headers = ["指标", "平均值", "最小值", "最大值", "中位数", "标准差", "P5", "P95"]
            ws_stats.append(stats_headers)

            for col, header in enumerate(stats_headers, 1):
                cell = ws_stats.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

            for data in assessments_data:
                stats = data["statistics"]
                if stats:
                    for metric in ["vmaf", "ssim", "psnr"]:
                        metric_stats = stats.get(metric)
                        if metric_stats:
                            ws_stats.append([
                                metric.upper(),
                                metric_stats.get("mean"),
                                metric_stats.get("min"),
                                metric_stats.get("max"),
                                metric_stats.get("median"),
                                metric_stats.get("std"),
                                metric_stats.get("p5"),
                                metric_stats.get("p95")
                            ])

        wb.save(output_path)

    async def _generate_pdf(
        self,
        output_path: Path,
        assessments_data: List[Dict],
        sections: List[str],
        report_name: str
    ) -> None:
        """生成 PDF 报告"""
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        styles = getSampleStyleSheet()
        story = []

        # 标题
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # 居中
        )
        story.append(Paragraph(report_name, title_style))
        story.append(Paragraph(
            f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            styles['Normal']
        ))
        story.append(Spacer(1, 20))

        # 摘要表格
        if "summary" in sections:
            story.append(Paragraph("评估摘要", styles['Heading2']))
            story.append(Spacer(1, 10))

            table_data = [["视频", "VMAF", "SSIM", "PSNR", "质量等级"]]

            for data in assessments_data:
                assessment = data["assessment"]
                vmaf = assessment.vmaf_score or 0

                if vmaf > 90:
                    level = "优秀"
                elif vmaf > 80:
                    level = "良好"
                elif vmaf > 70:
                    level = "可接受"
                else:
                    level = "差"

                table_data.append([
                    assessment.distorted_video.original_filename[:30],
                    f"{assessment.vmaf_score:.2f}" if assessment.vmaf_score else "N/A",
                    f"{assessment.ssim_score:.4f}" if assessment.ssim_score else "N/A",
                    f"{assessment.psnr_score:.2f}" if assessment.psnr_score else "N/A",
                    level
                ])

            table = Table(table_data, colWidths=[150, 60, 60, 60, 60])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            story.append(table)
            story.append(Spacer(1, 20))

        # 统计信息
        if "statistics" in sections:
            story.append(Paragraph("统计分析", styles['Heading2']))
            story.append(Spacer(1, 10))

            for data in assessments_data:
                stats = data["statistics"]
                assessment = data["assessment"]

                if stats:
                    story.append(Paragraph(
                        f"视频: {assessment.distorted_video.original_filename}",
                        styles['Heading3']
                    ))

                    for metric in ["vmaf", "ssim", "psnr"]:
                        metric_stats = stats.get(metric)
                        if metric_stats:
                            text = (
                                f"{metric.upper()}: "
                                f"平均={metric_stats.get('mean', 0):.2f}, "
                                f"最小={metric_stats.get('min', 0):.2f}, "
                                f"最大={metric_stats.get('max', 0):.2f}, "
                                f"标准差={metric_stats.get('std', 0):.2f}"
                            )
                            story.append(Paragraph(text, styles['Normal']))

                    story.append(Spacer(1, 10))

        doc.build(story)

    async def generate_share_link(
        self,
        session: AsyncSession,
        report_id: int,
        expires_days: int = 7
    ) -> str:
        """生成分享链接"""
        report = await session.get(Report, report_id)

        if not report:
            raise ValueError("报告不存在")

        # 生成唯一 token
        share_token = secrets.token_urlsafe(32)
        expires_at = datetime.utcnow() + timedelta(days=expires_days)

        report.share_token = share_token
        report.share_expires_at = expires_at

        await session.commit()

        return share_token

    async def get_report_by_token(
        self,
        session: AsyncSession,
        token: str
    ) -> Optional[Report]:
        """通过分享 token 获取报告"""
        query = select(Report).where(Report.share_token == token)
        result = await session.execute(query)
        report = result.scalar_one_or_none()

        if not report:
            return None

        # 检查是否过期
        if report.share_expires_at and report.share_expires_at < datetime.utcnow():
            return None

        return report

    async def list_reports(
        self,
        session: AsyncSession,
        skip: int = 0,
        limit: int = 20
    ) -> tuple[List[Report], int]:
        """获取报告列表"""
        count_query = select(Report)
        result = await session.execute(count_query)
        total = len(result.scalars().all())

        query = (
            select(Report)
            .order_by(Report.created_at.desc())
            .offset(skip)
            .limit(limit)
        )
        result = await session.execute(query)
        reports = result.scalars().all()

        return list(reports), total

    async def delete_report(
        self,
        session: AsyncSession,
        report_id: int
    ) -> bool:
        """删除报告"""
        report = await session.get(Report, report_id)

        if not report:
            return False

        # 删除报告文件
        report_dir = self.reports_dir / f"report_{report_id}"
        if report_dir.exists():
            import shutil
            shutil.rmtree(report_dir)

        await session.delete(report)
        await session.commit()

        return True

    async def create_batch_report(
        self,
        session: AsyncSession,
        batch_id: str,
        assessment_ids: List[int],
        reference_video: Video
    ) -> Report:
        """创建批量评估合并报告"""
        # 创建报告名称
        report_name = f"批量评估报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        # 创建报告记录
        report = Report(
            name=report_name,
            report_type="batch",
            assessment_ids={
                "ids": assessment_ids,
                "batch_id": batch_id,
                "sections": ["summary", "conclusion", "scatter", "charts", "statistics"]
            }
        )

        session.add(report)
        await session.commit()
        await session.refresh(report)

        # 生成报告文件
        await self._generate_batch_report_files(session, report, reference_video)

        return report

    async def _generate_batch_report_files(
        self,
        session: AsyncSession,
        report: Report,
        reference_video: Video
    ) -> None:
        """生成批量评估报告文件"""
        report_dir = self.reports_dir / f"report_{report.id}"
        report_dir.mkdir(parents=True, exist_ok=True)

        assessment_ids = report.assessment_ids.get("ids", [])

        # 获取评估数据
        assessments_data = []
        for aid in assessment_ids:
            query = (
                select(Assessment)
                .where(Assessment.id == aid)
                .options(
                    selectinload(Assessment.reference_video),
                    selectinload(Assessment.distorted_video)
                )
            )
            result = await session.execute(query)
            assessment = result.scalar_one_or_none()
            if assessment:
                frame_data = await assessment_service.get_frame_data(session, aid)
                stats = await assessment_service.get_statistics(session, aid)

                # 计算质量评级和效率
                vmaf = assessment.vmaf_score or 0
                bitrate_mbps = (assessment.distorted_video.bitrate or 0) / 1_000_000
                rating = get_quality_rating(vmaf)
                efficiency = calculate_efficiency(vmaf, bitrate_mbps)

                assessments_data.append({
                    "assessment": assessment,
                    "frame_data": frame_data,
                    "statistics": stats,
                    "rating": rating,
                    "efficiency": efficiency,
                    "bitrate_mbps": bitrate_mbps
                })

        # 按 VMAF 分数排序
        assessments_data.sort(key=lambda x: x["assessment"].vmaf_score or 0, reverse=True)

        # 生成 JSON
        json_path = report_dir / "report.json"
        await self._generate_batch_json(json_path, assessments_data, reference_video)
        report.json_path = str(json_path)

        # 生成 Excel
        excel_path = report_dir / "report.xlsx"
        await self._generate_batch_excel(excel_path, assessments_data, reference_video)
        report.excel_path = str(excel_path)

        # 生成散点图（三张并排 + 三张单独）
        scatter_path = report_dir / "scatter_chart.png"
        chart_paths = await self._generate_scatter_chart(scatter_path, assessments_data)

        # 生成 PDF
        pdf_path = report_dir / "report.pdf"
        await self._generate_batch_pdf(pdf_path, assessments_data, reference_video, report.name, scatter_path, chart_paths)
        report.pdf_path = str(pdf_path)

        await session.commit()

    async def _generate_scatter_chart(
        self,
        output_path: Path,
        assessments_data: List[Dict]
    ) -> Dict[str, Path]:
        """使用 matplotlib 生成三张散点图"""
        chart_paths = {}

        try:
            import matplotlib
            matplotlib.use('Agg')  # 使用非交互式后端
            import matplotlib.pyplot as plt

            # 尝试使用中文字体
            plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'PingFang SC', 'SimHei', 'DejaVu Sans']
            plt.rcParams['axes.unicode_minus'] = False

            # 提取数据
            bitrates = [d["bitrate_mbps"] for d in assessments_data]
            vmafs = [d["assessment"].vmaf_score or 0 for d in assessments_data]
            file_sizes = [(d["assessment"].distorted_video.file_size or 0) / 1_000_000 for d in assessments_data]  # MB
            names = [d["assessment"].distorted_video.original_filename[:15] for d in assessments_data]

            # 根据质量等级设置颜色
            def get_color(vmaf):
                if vmaf > 93:
                    return '#22c55e'  # green
                elif vmaf > 85:
                    return '#84cc16'  # lime
                elif vmaf > 70:
                    return '#eab308'  # yellow
                elif vmaf > 50:
                    return '#f97316'  # orange
                else:
                    return '#ef4444'  # red

            colors_list = [get_color(v) for v in vmafs]

            # === 生成三张并排的散点图 ===
            fig, axes = plt.subplots(1, 3, figsize=(18, 6))
            fig.suptitle('质量对比分析图', fontsize=16, fontweight='bold', y=1.02)

            # --- 左图：码率 vs 文件大小 ---
            ax1 = axes[0]
            ax1.scatter(bitrates, file_sizes, c=colors_list, s=120, alpha=0.8, edgecolors='white', linewidth=1.5)
            for i, name in enumerate(names):
                ax1.annotate(name, (bitrates[i], file_sizes[i]), textcoords="offset points",
                            xytext=(0, 8), ha='center', fontsize=7, alpha=0.8)
            ax1.set_xlabel('码率 (Mbps)', fontsize=11)
            ax1.set_ylabel('文件大小 (MB)', fontsize=11)
            ax1.set_title('码率 vs 文件大小', fontsize=12, fontweight='bold')
            ax1.grid(True, alpha=0.3)
            ax1.set_xlim(0, max(bitrates) * 1.1 if bitrates else 1)
            ax1.set_ylim(0, max(file_sizes) * 1.1 if file_sizes else 1)

            # --- 中图：码率 vs VMAF ---
            ax2 = axes[1]
            ax2.scatter(bitrates, vmafs, c=colors_list, s=120, alpha=0.8, edgecolors='white', linewidth=1.5)
            for i, name in enumerate(names):
                ax2.annotate(name, (bitrates[i], vmafs[i]), textcoords="offset points",
                            xytext=(0, 8), ha='center', fontsize=7, alpha=0.8)
            ax2.axhline(y=93, color='#22c55e', linestyle='--', alpha=0.6, linewidth=1)
            ax2.axhline(y=70, color='#eab308', linestyle='--', alpha=0.6, linewidth=1)
            ax2.axhspan(93, 100, alpha=0.08, color='#22c55e')
            ax2.axhspan(0, 70, alpha=0.08, color='#ef4444')
            ax2.set_xlabel('码率 (Mbps)', fontsize=11)
            ax2.set_ylabel('VMAF', fontsize=11)
            ax2.set_title('码率 vs VMAF', fontsize=12, fontweight='bold')
            ax2.grid(True, alpha=0.3)
            ax2.set_xlim(0, max(bitrates) * 1.1 if bitrates else 1)
            ax2.set_ylim(max(0, min(vmafs) - 5) if vmafs else 0, 100)

            # --- 右图：VMAF vs 文件大小 ---
            ax3 = axes[2]
            ax3.scatter(vmafs, file_sizes, c=colors_list, s=120, alpha=0.8, edgecolors='white', linewidth=1.5)
            for i, name in enumerate(names):
                ax3.annotate(name, (vmafs[i], file_sizes[i]), textcoords="offset points",
                            xytext=(0, 8), ha='center', fontsize=7, alpha=0.8)
            ax3.axvline(x=93, color='#22c55e', linestyle='--', alpha=0.6, linewidth=1)
            ax3.axvline(x=70, color='#eab308', linestyle='--', alpha=0.6, linewidth=1)
            ax3.set_xlabel('VMAF', fontsize=11)
            ax3.set_ylabel('文件大小 (MB)', fontsize=11)
            ax3.set_title('VMAF vs 文件大小', fontsize=12, fontweight='bold')
            ax3.grid(True, alpha=0.3)
            ax3.set_xlim(max(0, min(vmafs) - 5) if vmafs else 0, 100)
            ax3.set_ylim(0, max(file_sizes) * 1.1 if file_sizes else 1)

            # 添加图例
            from matplotlib.patches import Patch
            legend_elements = [
                Patch(facecolor='#22c55e', label='优秀 (>93)'),
                Patch(facecolor='#84cc16', label='良好 (85-93)'),
                Patch(facecolor='#eab308', label='可接受 (70-85)'),
                Patch(facecolor='#f97316', label='较差 (50-70)'),
                Patch(facecolor='#ef4444', label='很差 (<50)'),
            ]
            fig.legend(handles=legend_elements, loc='upper center', ncol=5,
                      bbox_to_anchor=(0.5, -0.02), fontsize=9)

            plt.tight_layout()
            plt.subplots_adjust(bottom=0.15)

            # 保存合并图
            plt.savefig(output_path, dpi=150, bbox_inches='tight', facecolor='white')
            chart_paths['combined'] = output_path
            plt.close()

            # === 生成单独的三张图片 ===
            output_dir = output_path.parent

            # 左图单独保存
            fig1, ax1 = plt.subplots(figsize=(8, 6))
            ax1.scatter(bitrates, file_sizes, c=colors_list, s=150, alpha=0.8, edgecolors='white', linewidth=2)
            for i, name in enumerate(names):
                ax1.annotate(name, (bitrates[i], file_sizes[i]), textcoords="offset points",
                            xytext=(0, 10), ha='center', fontsize=8)
            ax1.set_xlabel('码率 (Mbps)', fontsize=12)
            ax1.set_ylabel('文件大小 (MB)', fontsize=12)
            ax1.set_title('码率 vs 文件大小\n查看不同码率下文件大小的变化', fontsize=13, fontweight='bold')
            ax1.grid(True, alpha=0.3)
            ax1.set_xlim(0, max(bitrates) * 1.1 if bitrates else 1)
            ax1.set_ylim(0, max(file_sizes) * 1.1 if file_sizes else 1)
            plt.tight_layout()
            chart1_path = output_dir / "chart_bitrate_vs_size.png"
            plt.savefig(chart1_path, dpi=150, bbox_inches='tight', facecolor='white')
            chart_paths['bitrate_vs_size'] = chart1_path
            plt.close()

            # 中图单独保存
            fig2, ax2 = plt.subplots(figsize=(8, 6))
            ax2.scatter(bitrates, vmafs, c=colors_list, s=150, alpha=0.8, edgecolors='white', linewidth=2)
            for i, name in enumerate(names):
                ax2.annotate(name, (bitrates[i], vmafs[i]), textcoords="offset points",
                            xytext=(0, 10), ha='center', fontsize=8)
            ax2.axhline(y=93, color='#22c55e', linestyle='--', alpha=0.6, linewidth=1.5, label='优秀 (93)')
            ax2.axhline(y=70, color='#eab308', linestyle='--', alpha=0.6, linewidth=1.5, label='可接受 (70)')
            ax2.axhspan(93, 100, alpha=0.1, color='#22c55e')
            ax2.axhspan(0, 70, alpha=0.1, color='#ef4444')
            ax2.set_xlabel('码率 (Mbps)', fontsize=12)
            ax2.set_ylabel('VMAF', fontsize=12)
            ax2.set_title('码率 vs VMAF\n查看码率与画质之间的对应关系', fontsize=13, fontweight='bold')
            ax2.grid(True, alpha=0.3)
            ax2.legend(loc='lower right')
            ax2.set_xlim(0, max(bitrates) * 1.1 if bitrates else 1)
            ax2.set_ylim(max(0, min(vmafs) - 5) if vmafs else 0, 100)
            plt.tight_layout()
            chart2_path = output_dir / "chart_bitrate_vs_vmaf.png"
            plt.savefig(chart2_path, dpi=150, bbox_inches='tight', facecolor='white')
            chart_paths['bitrate_vs_vmaf'] = chart2_path
            plt.close()

            # 右图单独保存
            fig3, ax3 = plt.subplots(figsize=(8, 6))
            ax3.scatter(vmafs, file_sizes, c=colors_list, s=150, alpha=0.8, edgecolors='white', linewidth=2)
            for i, name in enumerate(names):
                ax3.annotate(name, (vmafs[i], file_sizes[i]), textcoords="offset points",
                            xytext=(0, 10), ha='center', fontsize=8)
            ax3.axvline(x=93, color='#22c55e', linestyle='--', alpha=0.6, linewidth=1.5, label='优秀 (93)')
            ax3.axvline(x=70, color='#eab308', linestyle='--', alpha=0.6, linewidth=1.5, label='可接受 (70)')
            ax3.set_xlabel('VMAF', fontsize=12)
            ax3.set_ylabel('文件大小 (MB)', fontsize=12)
            ax3.set_title('VMAF vs 文件大小\n查看画质提升带来的体积成本', fontsize=13, fontweight='bold')
            ax3.grid(True, alpha=0.3)
            ax3.legend(loc='upper left')
            ax3.set_xlim(max(0, min(vmafs) - 5) if vmafs else 0, 100)
            ax3.set_ylim(0, max(file_sizes) * 1.1 if file_sizes else 1)
            plt.tight_layout()
            chart3_path = output_dir / "chart_vmaf_vs_size.png"
            plt.savefig(chart3_path, dpi=150, bbox_inches='tight', facecolor='white')
            chart_paths['vmaf_vs_size'] = chart3_path
            plt.close()

        except ImportError:
            # 如果没有 matplotlib，跳过散点图生成
            pass

        return chart_paths

    async def _generate_batch_json(
        self,
        output_path: Path,
        assessments_data: List[Dict],
        reference_video: Video
    ) -> None:
        """生成批量评估 JSON 报告"""
        # 计算汇总数据
        total_count = len(assessments_data)
        avg_vmaf = sum(d["assessment"].vmaf_score or 0 for d in assessments_data) / total_count if total_count > 0 else 0
        best_video = assessments_data[0] if assessments_data else None
        best_efficiency = max(assessments_data, key=lambda x: x["efficiency"]["value"]) if assessments_data else None

        report_data = {
            "generated_at": datetime.utcnow().isoformat(),
            "report_type": "batch_comparison",
            "reference_video": {
                "filename": reference_video.original_filename,
                "resolution": f"{reference_video.width}x{reference_video.height}",
                "codec": reference_video.codec
            },
            "summary": {
                "total_assessments": total_count,
                "average_vmaf": round(avg_vmaf, 2),
                "best_quality_video": best_video["assessment"].distorted_video.original_filename if best_video else None,
                "best_efficiency_video": best_efficiency["assessment"].distorted_video.original_filename if best_efficiency else None
            },
            "conclusion_table": [],
            "assessments": []
        }

        # 生成结论表格数据
        for data in assessments_data:
            assessment = data["assessment"]
            rating = data["rating"]
            efficiency = data["efficiency"]

            report_data["conclusion_table"].append({
                "video_name": assessment.distorted_video.original_filename,
                "quality_stars": rating["stars"],
                "quality_level": rating["level"],
                "quality_description": rating["description"],
                "recommendation": rating["recommendation"],
                "recommendation_icon": rating["recommendation_icon"],
                "efficiency_level": efficiency["level"],
                "vmaf_score": assessment.vmaf_score,
                "bitrate_mbps": data["bitrate_mbps"]
            })

            # 详细评估数据
            report_data["assessments"].append({
                "id": assessment.id,
                "distorted_video": {
                    "filename": assessment.distorted_video.original_filename,
                    "resolution": f"{assessment.distorted_video.width}x{assessment.distorted_video.height}",
                    "codec": assessment.distorted_video.codec,
                    "bitrate": assessment.distorted_video.bitrate
                },
                "scores": {
                    "vmaf": assessment.vmaf_score,
                    "vmaf_min": assessment.vmaf_min,
                    "vmaf_max": assessment.vmaf_max,
                    "ssim": assessment.ssim_score,
                    "psnr": assessment.psnr_score
                },
                "rating": rating,
                "efficiency": efficiency,
                "statistics": data["statistics"]
            })

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)

    async def _generate_batch_excel(
        self,
        output_path: Path,
        assessments_data: List[Dict],
        reference_video: Video
    ) -> None:
        """生成批量评估 Excel 报告"""
        wb = Workbook()

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # === Sheet 1: 结论摘要 ===
        ws_conclusion = wb.active
        ws_conclusion.title = "结论摘要"

        # 标题行
        conclusion_headers = ["排名", "视频名称", "质量评级", "质量描述", "推荐程度", "码率效率", "VMAF", "码率(Mbps)"]
        ws_conclusion.append(conclusion_headers)

        for col, header in enumerate(conclusion_headers, 1):
            cell = ws_conclusion.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 数据行
        for rank, data in enumerate(assessments_data, 1):
            assessment = data["assessment"]
            rating = data["rating"]
            efficiency = data["efficiency"]

            row = [
                rank,
                assessment.distorted_video.original_filename,
                f"{rating['stars']} {rating['level']}",
                rating["description"],
                f"{rating['recommendation_icon']} {rating['recommendation']}",
                efficiency["level"],
                f"{assessment.vmaf_score:.2f}" if assessment.vmaf_score else "N/A",
                f"{data['bitrate_mbps']:.2f}"
            ]
            ws_conclusion.append(row)

            # 根据质量等级设置行颜色
            row_num = rank + 1
            vmaf = assessment.vmaf_score or 0
            if vmaf > 85:
                fill = green_fill
            elif vmaf > 70:
                fill = yellow_fill
            else:
                fill = red_fill

            for col in range(1, 9):
                cell = ws_conclusion.cell(row=row_num, column=col)
                cell.fill = fill
                cell.alignment = center_align
                cell.border = thin_border

        # 调整列宽
        ws_conclusion.column_dimensions['A'].width = 6
        ws_conclusion.column_dimensions['B'].width = 30
        ws_conclusion.column_dimensions['C'].width = 15
        ws_conclusion.column_dimensions['D'].width = 25
        ws_conclusion.column_dimensions['E'].width = 15
        ws_conclusion.column_dimensions['F'].width = 12
        ws_conclusion.column_dimensions['G'].width = 10
        ws_conclusion.column_dimensions['H'].width = 12

        # === Sheet 2: 详细数据 ===
        ws_detail = wb.create_sheet(title="详细数据")

        detail_headers = ["视频名称", "分辨率", "编码器", "码率(Mbps)", "VMAF", "VMAF Min", "VMAF Max", "SSIM", "PSNR"]
        ws_detail.append(detail_headers)

        for col, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for data in assessments_data:
            assessment = data["assessment"]
            dist_video = assessment.distorted_video

            row = [
                dist_video.original_filename,
                f"{dist_video.width}x{dist_video.height}",
                dist_video.codec or "N/A",
                f"{data['bitrate_mbps']:.2f}",
                f"{assessment.vmaf_score:.2f}" if assessment.vmaf_score else "N/A",
                f"{assessment.vmaf_min:.2f}" if assessment.vmaf_min else "N/A",
                f"{assessment.vmaf_max:.2f}" if assessment.vmaf_max else "N/A",
                f"{assessment.ssim_score:.4f}" if assessment.ssim_score else "N/A",
                f"{assessment.psnr_score:.2f}" if assessment.psnr_score else "N/A"
            ]
            ws_detail.append(row)

        # 调整列宽
        for col in ws_detail.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_detail.column_dimensions[col[0].column_letter].width = max_length + 2

        wb.save(output_path)

    async def _generate_batch_pdf(
        self,
        output_path: Path,
        assessments_data: List[Dict],
        reference_video: Video,
        report_name: str,
        scatter_path: Path,
        chart_paths: Dict[str, Path] = None
    ) -> None:
        """生成批量评估 PDF 报告"""
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        styles = getSampleStyleSheet()
        story = []

        # 标题样式
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            alignment=1,
            textColor=colors.gray
        )

        # === 标题页 ===
        story.append(Paragraph(report_name, title_style))
        story.append(Paragraph(
            f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            subtitle_style
        ))
        story.append(Paragraph(
            f"参考视频: {reference_video.original_filename}",
            subtitle_style
        ))
        story.append(Spacer(1, 30))

        # === 执行摘要 ===
        story.append(Paragraph("执行摘要", styles['Heading2']))
        story.append(Spacer(1, 10))

        # 汇总信息
        total_count = len(assessments_data)
        avg_vmaf = sum(d["assessment"].vmaf_score or 0 for d in assessments_data) / total_count if total_count > 0 else 0
        best_video = assessments_data[0] if assessments_data else None
        best_efficiency = max(assessments_data, key=lambda x: x["efficiency"]["value"]) if assessments_data else None

        summary_text = f"""
        本次批量评估共测试了 {total_count} 个视频文件。
        平均 VMAF 分数为 {avg_vmaf:.2f}。
        """
        if best_video:
            summary_text += f"最佳质量视频: {best_video['assessment'].distorted_video.original_filename}"
        if best_efficiency and best_efficiency != best_video:
            summary_text += f"\n最高性价比视频: {best_efficiency['assessment'].distorted_video.original_filename}"

        story.append(Paragraph(summary_text.strip(), styles['Normal']))
        story.append(Spacer(1, 20))

        # === 结论表格（易懂版） ===
        story.append(Paragraph("质量评估结论（简易版）", styles['Heading2']))
        story.append(Paragraph("以下表格帮助您快速了解各视频的质量情况", styles['Normal']))
        story.append(Spacer(1, 10))

        # 结论表格
        conclusion_data = [["排名", "视频", "质量", "推荐", "说明"]]

        for rank, data in enumerate(assessments_data[:10], 1):  # 最多显示 10 个
            assessment = data["assessment"]
            rating = data["rating"]

            conclusion_data.append([
                str(rank),
                assessment.distorted_video.original_filename[:25],
                f"{rating['stars']}\n{rating['level']}",
                f"{rating['recommendation_icon']}\n{rating['recommendation']}",
                rating["description"]
            ])

        conclusion_table = Table(conclusion_data, colWidths=[35, 120, 80, 70, 120])
        conclusion_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]))

        story.append(conclusion_table)
        story.append(Spacer(1, 20))

        # === 散点图 ===
        story.append(Paragraph("质量对比分析图", styles['Heading2']))
        story.append(Paragraph("以下三张图从不同维度展示各视频的质量、文件大小和 VMAF 分数之间的关系。", styles['Normal']))
        story.append(Spacer(1, 10))

        # 合并图（三张并排）
        if scatter_path.exists():
            img = Image(str(scatter_path), width=500, height=170)
            story.append(img)
            story.append(Spacer(1, 15))

        # 三张单独的图
        if chart_paths:
            chart_info = [
                ("bitrate_vs_size", "图1：码率 vs 文件大小", "斜率越陡表示编码效率越低，相同码率下文件更大"),
                ("bitrate_vs_vmaf", "图2：码率 vs VMAF", "曲线趋于平缓的位置是最佳码率点，再增加码率收益不大"),
                ("vmaf_vs_size", "图3：VMAF vs 文件大小", "越靠右下角的点性价比越高（高质量、小体积）"),
            ]

            for chart_key, title, desc in chart_info:
                if chart_key in chart_paths and chart_paths[chart_key].exists():
                    story.append(Paragraph(title, styles['Heading3']))
                    story.append(Paragraph(desc, styles['Normal']))
                    story.append(Spacer(1, 5))
                    img = Image(str(chart_paths[chart_key]), width=400, height=300)
                    story.append(img)
                    story.append(Spacer(1, 15))

        # === 详细数据表格 ===
        story.append(PageBreak())
        story.append(Paragraph("详细评估数据", styles['Heading2']))
        story.append(Spacer(1, 10))

        detail_data = [["视频", "VMAF", "SSIM", "PSNR", "码率", "编码"]]

        for data in assessments_data:
            assessment = data["assessment"]
            dist_video = assessment.distorted_video

            detail_data.append([
                dist_video.original_filename[:30],
                f"{assessment.vmaf_score:.2f}" if assessment.vmaf_score else "N/A",
                f"{assessment.ssim_score:.4f}" if assessment.ssim_score else "N/A",
                f"{assessment.psnr_score:.2f}" if assessment.psnr_score else "N/A",
                f"{data['bitrate_mbps']:.2f} Mbps",
                dist_video.codec or "N/A"
            ])

        detail_table = Table(detail_data, colWidths=[130, 55, 60, 55, 70, 50])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        story.append(detail_table)

        doc.build(story)


# 创建服务实例
report_service = ReportService()
